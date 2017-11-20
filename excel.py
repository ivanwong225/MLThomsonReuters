#Rohan Johar - CS 542
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook

#Workbook should have: today's price, boolean anomaly, headline, percent change, extraInfo
#Sample would be WB(5700, True, "The price has increased to 5700!", 25, "CHINA")
def WB(tPrice, anomaly, headline, pChange, extraInfo):
	if(os.path.isfile("BitcoinOutput.xlsx")):
		#Existing Workbook
		wb = load_workbook("BitcoinOutput.xlsx")
		ws = wb.active
		numberEntries = str(len(ws['A']) + 1)

		ws['A' + numberEntries] = tPrice
		if(anomaly):
			ws['B' + numberEntries] = 'Yes'
			ws['C' + numberEntries] = headline
			ws['E' + numberEntries] = extraInfo
		else:
			ws['B' + numberEntries] = 'No'
			ws['C' + numberEntries] = '-'
			ws['E' + numberEntries] = '-'
		ws['D' + numberEntries] = pChange
		
		try:
			wb.save('BitcoinOutput.xlsx')
		except PermissionError:
			return
	else:
		#New Workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Anomaly Output"
		ws['A1'] = "Today's Price"
		ws['B1'] = "Anomaly?"
		ws['C1'] = "Headline"
		ws['D1'] = "Percent Change"
		ws['E1'] = "Extra information"

		ws['A2'] = tPrice
		if(anomaly):
			ws['B2'] = 'Yes'
			ws['C2'] = headline
			ws['E2'] = extraInfo
		else:
			ws['B2'] = 'No'
			ws['C2'] = '-'
			ws['E2'] = '-'
		ws['D2'] = pChange

		wb.save('BitcoinOutput.xlsx')